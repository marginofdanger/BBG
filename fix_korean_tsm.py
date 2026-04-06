"""Add SK Hynix/Samsung, fix TSM ADR currency in comp sheet."""
from openpyxl import load_workbook
from openpyxl.styles import Font
from xbbg import blp

OUT = r"C:\Users\AdrianOw\Projects\BBG\output\comp_sheet_updated.xlsx"
SRC = r"C:\Users\AdrianOw\OneDrive - chieftaincapital.com\comp sheet.xlsx"
MONTH_NAMES = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
               7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
USD_OVERRIDE = {"TSM US Equity", "TSM UN Equity"}

def bdp_d(tickers, fields):
    result = {}
    for i in range(0, len(tickers), 40):
        try:
            df = blp.bdp(tickers[i:i+40], fields)
            for r in df.rows():
                if r[0] not in result: result[r[0]] = {}
                try: result[r[0]][r[1]] = float(r[2])
                except: result[r[0]][r[1]] = r[2]
        except: pass
    return result

def bdp_ov(tickers, field, period, extra_overrides=None):
    result = {}
    ovr = [("BEST_FPERIOD_OVERRIDE", period)]
    if extra_overrides:
        ovr.extend(extra_overrides)
    try:
        df = blp.bdp(tickers, field, overrides=ovr)
        for r in df.rows():
            if r[0] not in result: result[r[0]] = {}
            try: result[r[0]][r[1]] = float(r[2])
            except: result[r[0]][r[1]] = r[2]
    except: pass
    return result

def main():
    wb_src = load_workbook(SRC, data_only=True)
    ws_src = wb_src.active
    src_map = {}
    for row in range(1, ws_src.max_row + 1):
        t = ws_src.cell(row, 1).value
        b = ws_src.cell(row, 2).value
        if t and b and "Equity" in str(b):
            src_map[row] = {"ticker": str(t), "bbg": str(b)}

    wb = load_workbook(OUT)
    ws = wb["Data"]

    # Check existing tickers
    existing = set()
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row, 1).value
        if v: existing.add(str(v))

    # Add Korean stocks
    insert_row = ws.max_row + 1
    new_kr = [
        ("000660", "000660 KS Equity", "SK Hynix"),
        ("005930", "005930 KS Equity", "Samsung"),
    ]
    added = []
    need_header = True
    for ticker, bbg, name in new_kr:
        if ticker in existing:
            print(f"SKIP {ticker}")
            continue
        if need_header:
            ws.cell(insert_row, 1).value = "Memory (Korea)"
            ws.cell(insert_row, 1).font = Font(bold=True)
            insert_row += 1
            need_header = False
        ws.cell(insert_row, 1).value = ticker
        ws.cell(insert_row, 2).value = bbg
        ws.cell(insert_row, 3).value = name
        src_map[insert_row] = {"ticker": ticker, "bbg": bbg}
        added.append({"row": insert_row, "ticker": ticker, "bbg": bbg})
        print(f"ADD {ticker} ({name}) at row {insert_row}")
        insert_row += 1

    # Pull Korean data
    if added:
        bbg_list = [a["bbg"] for a in added]
        snap = bdp_d(bbg_list, ["PX_LAST","CUR_MKT_CAP","CHG_PCT_YTD","NET_DEBT_TO_EBITDA","RETURN_ON_INV_CAPITAL","TRAIL_12M_EBITDA"])
        eps, rev, ebit = {}, {}, {}
        for yr in ["24","25","26","27","28"]:
            for field, store, key in [("BEST_EPS",eps,"EPS"),("BEST_SALES",rev,"REV"),("BEST_EBIT",ebit,"EBIT")]:
                d = bdp_ov(bbg_list, field, f"{yr}Y")
                for t in d:
                    if t not in store: store[t] = {}
                    store[t][f"{key}_20{yr}"] = d[t].get(field)

        for a in added:
            row, bbg = a["row"], a["bbg"]
            s = snap.get(bbg, {})
            e = eps.get(bbg, {})
            r = rev.get(bbg, {})
            price = s.get("PX_LAST")
            mktcap = s.get("CUR_MKT_CAP")
            if mktcap: mktcap /= 1e9
            if price: ws.cell(row, 5).value = price
            if mktcap: ws.cell(row, 4).value = mktcap
            for ci, yr in enumerate(["2026","2027","2028"]):
                ev = e.get(f"EPS_{yr}")
                if price and ev and ev != 0:
                    ws.cell(row, 6+ci).value = price / ev
            net_levg = s.get("NET_DEBT_TO_EBITDA")
            if net_levg: ws.cell(row, 9).value = net_levg
            roic = s.get("RETURN_ON_INV_CAPITAL")
            if roic: ws.cell(row, 12).value = roic/100
            ytd = s.get("CHG_PCT_YTD")
            if ytd: ws.cell(row, 17).value = ytd/100
            for ci, yr in enumerate(["2024","2025","2026","2027","2028"]):
                v = e.get(f"EPS_{yr}")
                if v: ws.cell(row, 36+ci).value = v
                v = r.get(f"REV_{yr}")
                if v: ws.cell(row, 42+ci).value = v
            print(f"  {a['ticker']}: Price={price}, PE26={ws.cell(row,6).value}")

    # Fix TSM PE: pull EPS in USD
    print("\nFixing TSM PE...")
    for row in range(1, ws.max_row + 1):
        bbg_cell = str(ws.cell(row, 2).value or "")
        if "TSM" not in bbg_cell or "Equity" not in bbg_cell:
            continue
        price = ws.cell(row, 5).value
        if not price: continue
        price = float(price)
        for ci, yr in enumerate(["26","27","28"]):
            d = bdp_ov(["TSM US Equity"], "BEST_EPS", f"{yr}Y", [("EQY_FUND_CRNCY","USD")])
            eps_usd = d.get("TSM US Equity", {}).get("BEST_EPS")
            if eps_usd and eps_usd != 0:
                ws.cell(row, 6+ci).value = price / eps_usd
                print(f"  Row {row} PE {yr}: {price/eps_usd:.1f}x")
        for ci, yr in enumerate(["24","25","26","27","28"]):
            d = bdp_ov(["TSM US Equity"], "BEST_EPS", f"{yr}Y", [("EQY_FUND_CRNCY","USD")])
            eps_usd = d.get("TSM US Equity", {}).get("BEST_EPS")
            if eps_usd: ws.cell(row, 36+ci).value = eps_usd
        # Revenue in USD too
        for ci, yr in enumerate(["24","25","26","27","28"]):
            d = bdp_ov(["TSM US Equity"], "BEST_SALES", f"{yr}Y", [("EQY_FUND_CRNCY","USD")])
            rev_usd = d.get("TSM US Equity", {}).get("BEST_SALES")
            if rev_usd: ws.cell(row, 42+ci).value = rev_usd

    # Rebuild BBG Live
    print("\nRebuilding BBG Live...")
    if "BBG Live" in wb.sheetnames:
        del wb["BBG Live"]
    live = wb.copy_worksheet(ws)
    live.title = "BBG Live"

    all_bbg = list(set(info["bbg"] for info in src_map.values()))
    fye_data = bdp_d(all_bbg, ["EQY_FISCAL_YR_END"])
    fye_flat = {bbg: fields.get("EQY_FISCAL_YR_END", "") for bbg, fields in fye_data.items()}
    blue = Font(color="0000FF")

    live.cell(2, 3).value = "FYE"
    live.cell(1, 6).value = "PE"
    live.cell(2, 6).value = 2026
    live.cell(2, 7).value = 2027
    live.cell(2, 8).value = 2028
    live.cell(1, 10).value = "2026 estimates"
    live.cell(1, 13).value = "Growth: 2025-2028"
    live.cell(2, 20).value = "25-28 EPS gwth"

    for row, info in src_map.items():
        bbg = info["bbg"]
        raw = fye_flat.get(bbg, "")
        month = 12
        if raw and "/" in str(raw):
            try: month = int(str(raw).split("/")[0])
            except: month = 12
        fye_label = MONTH_NAMES.get(month, "Dec")
        offset = 1 if month <= 3 else 0
        needs_usd = bbg in USD_OVERRIDE
        live.cell(row, 3).value = fye_label

        def fy(cy):
            return f"{(cy+offset)%100}Y"

        def bdp(fld, ov=None):
            if needs_usd and fld in ("BEST_EPS","BEST_SALES","BEST_EBIT") and ov:
                return f'=BDP("{bbg}","{fld}","BEST_FPERIOD_OVERRIDE","{ov}","EQY_FUND_CRNCY","USD")'
            if ov:
                return f'=BDP("{bbg}","{fld}","BEST_FPERIOD_OVERRIDE","{ov}")'
            return f'=BDP("{bbg}","{fld}")'

        pr = f"$E${row}"
        live.cell(row, 4).value = f"={bdp('CUR_MKT_CAP')[1:]}/1000000000"
        live.cell(row, 5).value = bdp("PX_LAST")
        for ci, cy in enumerate([2026,2027,2028]):
            live.cell(row,6+ci).value = f'=IF({bdp("BEST_EPS",fy(cy))[1:]}<>0,{pr}/{bdp("BEST_EPS",fy(cy))[1:]},"")'
        live.cell(row, 9).value = bdp("NET_DEBT_TO_EBITDA")
        live.cell(row, 10).value = f"={bdp('BEST_SALES',fy(2026))[1:]}/1000"
        live.cell(row, 11).value = f'=IF({bdp("BEST_SALES",fy(2026))[1:]}<>0,{bdp("BEST_EBIT",fy(2026))[1:]}/{bdp("BEST_SALES",fy(2026))[1:]},"")'
        live.cell(row, 12).value = f"={bdp('RETURN_ON_INV_CAPITAL')[1:]}/100"
        for col, fld in [(13,"BEST_SALES"),(14,"BEST_EBIT"),(15,"BEST_EPS")]:
            live.cell(row, col).value = f'=IF(AND({bdp(fld,fy(2025))[1:]}>0,{bdp(fld,fy(2028))[1:]}>0),({bdp(fld,fy(2028))[1:]}/{bdp(fld,fy(2025))[1:]})^(1/3)-1,"")'
        live.cell(row, 17).value = f"={bdp('CHG_PCT_YTD')[1:]}/100"
        live.cell(row, 20).value = f'=IF(AND(F{row}<>"",O{row}<>"",O{row}<>0),F{row}/(O{row}*100),"")'
        live.cell(row, 28).value = bdp("TRAIL_12M_OPER_INC")
        live.cell(row, 29).value = bdp("SHORT_AND_LONG_TERM_DEBT")
        live.cell(row, 30).value = bdp("TOTAL_EQUITY")
        live.cell(row, 31).value = bdp("BS_CASH_NEAR_CASH_ITEM")
        live.cell(row, 32).value = bdp("BS_LT_INVEST")
        live.cell(row, 33).value = f"=AC{row}+AD{row}"
        live.cell(row, 34).value = f"={bdp('RETURN_ON_INV_CAPITAL')[1:]}/100"
        for ci, cy in enumerate([2024,2025,2026,2027,2028]):
            live.cell(row, 36+ci).value = bdp("BEST_EPS", fy(cy))
            live.cell(row, 42+ci).value = bdp("BEST_SALES", fy(cy))
            live.cell(row, 48+ci).value = bdp("BEST_EBIT", fy(cy))
        live.cell(row, 54).value = f"=D{row}+BC{row}"
        live.cell(row, 55).value = f"=AC{row}-AE{row}"
        live.cell(row, 56).value = f"=D{row}"
        live.cell(row, 57).value = bdp("TRAIL_12M_EBITDA")

    for row in range(1, live.max_row + 1):
        for col in range(1, live.max_column + 1):
            val = live.cell(row, col).value
            if val and isinstance(val, str) and "BDP(" in val:
                live.cell(row, col).font = blue

    wb.save(OUT)
    print(f"\nSaved: {OUT}")

if __name__ == "__main__":
    main()

"""Build comp sheet with BBG Live formulas. Single source of truth."""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from xbbg import blp

SRC = r"C:\Users\AdrianOw\OneDrive - chieftaincapital.com\comp sheet.xlsx"
OUT = r"C:\Users\AdrianOw\Projects\BBG\output\comp_sheet_updated.xlsx"

MONTH_NAMES = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
               7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}

# Tickers needing USD currency override (ADRs where price is USD but estimates default to local)
USD_OVERRIDE = {"TSM US Equity", "TSM UN Equity"}

# Korean tickers: price/EPS in KRW, mktcap/rev divide by 1e6 to get trillions KRW
KRW_TICKERS = {"000660 KS Equity", "005930 KS Equity"}

# Additional tickers to integrate (not in source file)
EXTRA_TICKERS = [
    (None, "Additional", None),  # Category header
    ("CRM", "CRM US Equity", "Salesforce"),
    ("ORCL", "ORCL US Equity", "Oracle"),
    ("CME", "CME US Equity", "CME Group"),
    ("ICE", "ICE US Equity", "ICE/NYSE"),
    ("IR", "IR US Equity", "Ingersoll Rand"),
    ("PANW", "PANW US Equity", "Palo Alto Networks"),
    (None, "Memory (Korea)", None),  # Category header
    ("000660", "000660 KS Equity", "SK Hynix"),
    ("005930", "005930 KS Equity", "Samsung"),
]


def bdp_d(tickers, fields):
    result = {}
    for i in range(0, len(tickers), 40):
        try:
            df = blp.bdp(tickers[i:i+40], fields)
            for r in df.rows():
                if r[0] not in result: result[r[0]] = {}
                try: result[r[0]][r[1]] = float(r[2])
                except: result[r[0]][r[1]] = r[2]
        except: pass
    return result


def bdp_ov(tickers, field, period, extra=None):
    result = {}
    ovr = [("BEST_FPERIOD_OVERRIDE", period)]
    if extra: ovr.extend(extra)
    for i in range(0, len(tickers), 40):
        try:
            df = blp.bdp(tickers[i:i+40], field, overrides=ovr)
            for r in df.rows():
                if r[0] not in result: result[r[0]] = {}
                try: result[r[0]][r[1]] = float(r[2])
                except: result[r[0]][r[1]] = r[2]
        except: pass
    return result


def main():
    # Load source
    wb_src = load_workbook(SRC, data_only=True)
    ws_src = wb_src.active
    wb = load_workbook(SRC)
    ws = wb.active

    # Build ticker map from source
    src_map = {}
    existing = set()
    for row in range(1, ws_src.max_row + 1):
        t = ws_src.cell(row, 1).value
        b = ws_src.cell(row, 2).value
        if t and b and "Equity" in str(b):
            src_map[row] = {"ticker": str(t), "bbg": str(b)}
            existing.add(str(t))

    # Add extra tickers
    insert_row = ws.max_row + 2
    for ticker, bbg_or_cat, name in EXTRA_TICKERS:
        if ticker is None:
            # Category header
            ws.cell(insert_row, 1).value = bbg_or_cat
            ws.cell(insert_row, 1).font = Font(bold=True)
            insert_row += 1
            continue
        if ticker in existing:
            continue
        ws.cell(insert_row, 1).value = ticker
        ws.cell(insert_row, 2).value = bbg_or_cat
        if name: ws.cell(insert_row, 3).value = name
        src_map[insert_row] = {"ticker": ticker, "bbg": bbg_or_cat}
        existing.add(ticker)
        print(f"ADD {ticker} ({name})")
        insert_row += 1

    all_info = list(src_map.items())
    all_bbg = list(set(info["bbg"] for _, info in all_info))
    print(f"\nTotal: {len(all_info)} tickers, {len(all_bbg)} unique")

    # Pull all data
    print("\n[1/3] Pulling snapshot data...")
    snap = bdp_d(all_bbg, ["PX_LAST","CUR_MKT_CAP","CHG_PCT_YTD","NET_DEBT_TO_EBITDA",
        "RETURN_ON_INV_CAPITAL","SHORT_AND_LONG_TERM_DEBT","BS_CASH_NEAR_CASH_ITEM",
        "TOTAL_EQUITY","BS_LT_INVEST","TRAIL_12M_OPER_INC","TRAIL_12M_EBITDA"])

    print("\n[2/3] Pulling estimates 2024-2028...")
    eps, rev, ebit = {}, {}, {}
    for yr in ["24","25","26","27","28"]:
        for field, store, key in [("BEST_EPS",eps,"EPS"),("BEST_SALES",rev,"REV"),("BEST_EBIT",ebit,"EBIT")]:
            print(f"  [{key}] {yr}Y...")
            # Standard pull
            d = bdp_ov(all_bbg, field, f"{yr}Y")
            for t in d:
                if t not in store: store[t] = {}
                store[t][f"{key}_20{yr}"] = d[t].get(field)
            # USD override for ADR tickers
            usd_tickers = [t for t in all_bbg if t in USD_OVERRIDE]
            if usd_tickers:
                d2 = bdp_ov(usd_tickers, field, f"{yr}Y", [("EQY_FUND_CRNCY","USD")])
                for t in d2:
                    if t not in store: store[t] = {}
                    store[t][f"{key}_20{yr}"] = d2[t].get(field)

    print("\n[3/3] Pulling FYE...")
    fye_data = bdp_d(all_bbg, ["EQY_FISCAL_YR_END"])
    fye_flat = {bbg: fields.get("EQY_FISCAL_YR_END", "") for bbg, fields in fye_data.items()}

    def get_offset(bbg):
        raw = fye_flat.get(bbg, "")
        month = 12
        if raw and "/" in str(raw):
            try: month = int(str(raw).split("/")[0])
            except: month = 12
        return MONTH_NAMES.get(month, "Dec"), 1 if month <= 3 else 0

    # --- Update Data sheet ---
    print("\nUpdating Data sheet...")
    ws.title = "Data"
    ws.cell(1, 6).value = "PE"
    ws.cell(2, 6).value = 2026
    ws.cell(2, 7).value = 2027
    ws.cell(2, 8).value = 2028
    ws.cell(1, 10).value = "2026 estimates"
    ws.cell(1, 13).value = "Growth: 2025-2028"
    ws.cell(2, 20).value = "25-28 EPS gwth"

    for row, info in all_info:
        bbg = info["bbg"]
        s = snap.get(bbg, {})
        e = eps.get(bbg, {})
        r = rev.get(bbg, {})
        eb = ebit.get(bbg, {})
        is_krw = bbg in KRW_TICKERS

        price = s.get("PX_LAST")
        mktcap = s.get("CUR_MKT_CAP")
        # Mktcap: billions USD, or trillions KRW
        if mktcap:
            mktcap = mktcap / 1e12 if is_krw else mktcap / 1e9

        if price: ws.cell(row, 5).value = price
        if mktcap: ws.cell(row, 4).value = mktcap

        eps_26 = e.get("EPS_2026")
        eps_27 = e.get("EPS_2027")
        eps_28 = e.get("EPS_2028")
        if price and eps_26 and eps_26 != 0: ws.cell(row, 6).value = price/eps_26
        if price and eps_27 and eps_27 != 0: ws.cell(row, 7).value = price/eps_27
        if price and eps_28 and eps_28 != 0: ws.cell(row, 8).value = price/eps_28

        net_levg = s.get("NET_DEBT_TO_EBITDA")
        if net_levg: ws.cell(row, 9).value = net_levg

        rev_26 = r.get("REV_2026")
        ebit_26 = eb.get("EBIT_2026")
        # Revenue: billions USD, or trillions KRW
        if rev_26:
            ws.cell(row, 10).value = rev_26 / 1e6 if is_krw else rev_26 / 1000
        if rev_26 and ebit_26 and rev_26 != 0:
            ws.cell(row, 11).value = ebit_26 / rev_26
        roic = s.get("RETURN_ON_INV_CAPITAL")
        if roic: ws.cell(row, 12).value = roic/100

        def cagr(start, end, yrs=3):
            if not start or not end or start <= 0 or end <= 0: return None
            return (end/start)**(1/yrs)-1

        v = cagr(r.get("REV_2025"), r.get("REV_2028"))
        if v: ws.cell(row, 13).value = v
        v = cagr(eb.get("EBIT_2025"), eb.get("EBIT_2028"))
        if v: ws.cell(row, 14).value = v
        v = cagr(e.get("EPS_2025"), e.get("EPS_2028"))
        if v: ws.cell(row, 15).value = v

        ytd = s.get("CHG_PCT_YTD")
        if ytd: ws.cell(row, 17).value = ytd/100

        try:
            pe26 = float(ws.cell(row, 6).value) if ws.cell(row, 6).value else None
            eg = float(ws.cell(row, 15).value) if ws.cell(row, 15).value else None
            if pe26 and eg and eg != 0:
                ws.cell(row, 20).value = pe26/(eg*100)
        except: pass

        ltm_opinc = s.get("TRAIL_12M_OPER_INC")
        if ltm_opinc: ws.cell(row, 28).value = ltm_opinc
        debt = s.get("SHORT_AND_LONG_TERM_DEBT")
        equity = s.get("TOTAL_EQUITY")
        cash = s.get("BS_CASH_NEAR_CASH_ITEM")
        lti = s.get("BS_LT_INVEST")
        if debt: ws.cell(row, 29).value = debt
        if equity: ws.cell(row, 30).value = equity
        if cash: ws.cell(row, 31).value = cash
        if lti: ws.cell(row, 32).value = lti
        if debt and equity: ws.cell(row, 33).value = debt + equity
        if roic: ws.cell(row, 34).value = roic/100

        for ci, yr in enumerate(["2024","2025","2026","2027","2028"]):
            v = e.get(f"EPS_{yr}")
            if v: ws.cell(row, 36+ci).value = v
            v = r.get(f"REV_{yr}")
            if v: ws.cell(row, 42+ci).value = v
            v = eb.get(f"EBIT_{yr}")
            if v: ws.cell(row, 48+ci).value = v

        if mktcap:
            ws.cell(row, 56).value = mktcap
            net_debt = (debt or 0) - (cash or 0)
            ws.cell(row, 55).value = net_debt
            ws.cell(row, 54).value = mktcap + net_debt
        ltm_ebitda = s.get("TRAIL_12M_EBITDA")
        if ltm_ebitda: ws.cell(row, 57).value = ltm_ebitda

    # --- Build BBG Live sheet ---
    print("Building BBG Live...")
    if "BBG Live" in wb.sheetnames:
        del wb["BBG Live"]
    live = wb.copy_worksheet(ws)
    live.title = "BBG Live"

    blue = Font(color="0000FF")
    live.cell(2, 3).value = "FYE"
    live.cell(1, 6).value = "PE"
    live.cell(2, 6).value = 2026
    live.cell(2, 7).value = 2027
    live.cell(2, 8).value = 2028
    live.cell(1, 10).value = "2026 estimates"
    live.cell(1, 13).value = "Growth: 2025-2028"
    live.cell(2, 20).value = "25-28 EPS gwth"

    for row, info in all_info:
        bbg = info["bbg"]
        fye_label, offset = get_offset(bbg)
        needs_usd = bbg in USD_OVERRIDE
        is_krw = bbg in KRW_TICKERS
        live.cell(row, 3).value = fye_label

        def fy(cy): return f"{(cy+offset)%100}Y"
        def bdp(fld, ov=None):
            if (needs_usd) and fld in ("BEST_EPS","BEST_SALES","BEST_EBIT") and ov:
                return f'=BDP("{bbg}","{fld}","BEST_FPERIOD_OVERRIDE","{ov}","EQY_FUND_CRNCY","USD")'
            if ov:
                return f'=BDP("{bbg}","{fld}","BEST_FPERIOD_OVERRIDE","{ov}")'
            return f'=BDP("{bbg}","{fld}")'

        # Mktcap divisor: 1e12 for KRW (trillions), 1e9 for USD (billions)
        mktcap_div = "1000000000000" if is_krw else "1000000000"
        rev_div = "1000000" if is_krw else "1000"

        pr = f"$E${row}"
        live.cell(row, 4).value = f"={bdp('CUR_MKT_CAP')[1:]}/{mktcap_div}"
        live.cell(row, 5).value = bdp("PX_LAST")
        for ci, cy in enumerate([2026,2027,2028]):
            live.cell(row,6+ci).value = f'=IF({bdp("BEST_EPS",fy(cy))[1:]}<>0,{pr}/{bdp("BEST_EPS",fy(cy))[1:]},"")'
        live.cell(row, 9).value = bdp("NET_DEBT_TO_EBITDA")
        live.cell(row, 10).value = f"={bdp('BEST_SALES',fy(2026))[1:]}/{rev_div}"
        live.cell(row, 11).value = f'=IF({bdp("BEST_SALES",fy(2026))[1:]}<>0,{bdp("BEST_EBIT",fy(2026))[1:]}/{bdp("BEST_SALES",fy(2026))[1:]},"")'
        live.cell(row, 12).value = f"={bdp('RETURN_ON_INV_CAPITAL')[1:]}/100"
        for col, fld in [(13,"BEST_SALES"),(14,"BEST_EBIT"),(15,"BEST_EPS")]:
            live.cell(row, col).value = f'=IF(AND({bdp(fld,fy(2025))[1:]}>0,{bdp(fld,fy(2028))[1:]}>0),({bdp(fld,fy(2028))[1:]}/{bdp(fld,fy(2025))[1:]})^(1/3)-1,"")'
        live.cell(row, 17).value = f"={bdp('CHG_PCT_YTD')[1:]}/100"
        live.cell(row, 20).value = f'=IF(AND(F{row}<>"",O{row}<>"",O{row}<>0),F{row}/(O{row}*100),"")'
        live.cell(row, 28).value = bdp("TRAIL_12M_OPER_INC")
        live.cell(row, 29).value = bdp("SHORT_AND_LONG_TERM_DEBT")
        live.cell(row, 30).value = bdp("TOTAL_EQUITY")
        live.cell(row, 31).value = bdp("BS_CASH_NEAR_CASH_ITEM")
        live.cell(row, 32).value = bdp("BS_LT_INVEST")
        live.cell(row, 33).value = f"=AC{row}+AD{row}"
        live.cell(row, 34).value = f"={bdp('RETURN_ON_INV_CAPITAL')[1:]}/100"
        for ci, cy in enumerate([2024,2025,2026,2027,2028]):
            live.cell(row, 36+ci).value = bdp("BEST_EPS", fy(cy))
            live.cell(row, 42+ci).value = bdp("BEST_SALES", fy(cy))
            live.cell(row, 48+ci).value = bdp("BEST_EBIT", fy(cy))
        live.cell(row, 54).value = f"=D{row}+BC{row}"
        live.cell(row, 55).value = f"=AC{row}-AE{row}"
        live.cell(row, 56).value = f"=D{row}"
        live.cell(row, 57).value = bdp("TRAIL_12M_EBITDA")

    for row in range(1, live.max_row + 1):
        for col in range(1, live.max_column + 1):
            val = live.cell(row, col).value
            if val and isinstance(val, str) and "BDP(" in val:
                live.cell(row, col).font = blue

    # --- Build BBG Formulas reference sheet ---
    print("Building BBG Formulas...")
    if "BBG Formulas" in wb.sheetnames:
        del wb["BBG Formulas"]
    fs = wb.create_sheet("BBG Formulas")
    fs.cell(1,1).value = "Ticker"
    fs.cell(1,2).value = "BBG Ticker"
    fs.cell(1,3).value = "Field"
    fs.cell(1,4).value = "Override"
    fs.cell(1,5).value = "Description"
    fs.cell(1,6).value = "BDP Formula"
    for c in range(1,7): fs.cell(1,c).font = Font(bold=True)

    frow = 2
    seen = set()
    for _, info in all_info:
        bbg = info["bbg"]
        if bbg in seen: continue
        seen.add(bbg)
        t = info["ticker"]
        formulas = [("PX_LAST","","Price"),("CUR_MKT_CAP","","Mkt cap"),
            ("CHG_PCT_YTD","","YTD chg"),("NET_DEBT_TO_EBITDA","","Net leverage"),
            ("RETURN_ON_INV_CAPITAL","","ROIC"),("TRAIL_12M_OPER_INC","","LTM Op inc"),
            ("TRAIL_12M_EBITDA","","LTM EBITDA"),("SHORT_AND_LONG_TERM_DEBT","","Debt"),
            ("TOTAL_EQUITY","","Equity"),("BS_CASH_NEAR_CASH_ITEM","","Cash")]
        for yr in ["24","25","26","27","28"]:
            formulas += [(f,"BEST_FPERIOD_OVERRIDE={yr}Y",f"FY20{yr}")
                for f in ["BEST_EPS","BEST_SALES","BEST_EBIT"]]
        for field, ov, desc in formulas:
            fs.cell(frow,1).value = t
            fs.cell(frow,2).value = bbg
            fs.cell(frow,3).value = field
            fs.cell(frow,4).value = ov
            fs.cell(frow,5).value = desc
            frow += 1

    for col in range(1,7): fs.column_dimensions[get_column_letter(col)].width = 25

    # Remove Data sheet - user only wants BBG Live
    if "Data" in wb.sheetnames:
        del wb["Data"]

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"\nSaved: {OUT}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()

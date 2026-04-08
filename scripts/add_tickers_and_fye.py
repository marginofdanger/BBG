"""Post-process: add new tickers to comp sheet and rebuild BBG Live with FYE-adjusted formulas."""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from xbbg import blp

OUT = r"C:\Users\AdrianOw\Projects\BBG\output\comp_sheet_updated.xlsx"
SRC = r"C:\Users\AdrianOw\OneDrive - chieftaincapital.com\comp sheet.xlsx"

MONTH_NAMES = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
               7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}

NEW_TICKERS = [
    ('CRM', 'CRM US Equity'),
    ('ORCL', 'ORCL US Equity'),
    ('CME', 'CME US Equity'),
    ('ICE', 'ICE US Equity'),
    ('IR', 'IR US Equity'),
    ('PANW', 'PANW US Equity'),
]


def bdp_dict(tickers, fields):
    result = {}
    for i in range(0, len(tickers), 40):
        try:
            df = blp.bdp(tickers[i:i+40], fields)
            for r in df.rows():
                if r[0] not in result: result[r[0]] = {}
                try: result[r[0]][r[1]] = float(r[2])
                except: result[r[0]][r[1]] = r[2]
        except Exception as e:
            print(f"  WARNING: {e}")
    return result


def bdp_ov(tickers, field, period):
    result = {}
    for i in range(0, len(tickers), 40):
        try:
            df = blp.bdp(tickers[i:i+40], field, overrides=[('BEST_FPERIOD_OVERRIDE', period)])
            for r in df.rows():
                if r[0] not in result: result[r[0]] = {}
                try: result[r[0]][r[1]] = float(r[2])
                except: result[r[0]][r[1]] = r[2]
        except: pass
    return result


def get_fye_offset(bbg, fye_data):
    raw = fye_data.get(bbg, '')
    month = 12
    if raw and '/' in str(raw):
        try: month = int(str(raw).split('/')[0])
        except: month = 12
    return MONTH_NAMES.get(month, 'Dec'), 1 if month <= 3 else 0


def main():
    # Get resolved tickers from source
    wb_src = load_workbook(SRC, data_only=True)
    ws_src = wb_src.active
    src_map = {}
    for row in range(1, ws_src.max_row + 1):
        t = ws_src.cell(row, 1).value
        b = ws_src.cell(row, 2).value
        if t and b and 'Equity' in str(b):
            src_map[row] = {'ticker': str(t), 'bbg': str(b)}

    # Load output
    wb = load_workbook(OUT)
    ws = wb['Data']

    # --- Add new tickers ---
    existing = set()
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row, 1).value
        if v: existing.add(str(v))

    insert_row = ws.max_row + 2
    added = []
    need_header = True

    for ticker, bbg in NEW_TICKERS:
        if ticker in existing:
            print(f"SKIP {ticker}")
            continue
        if need_header:
            ws.cell(insert_row, 1).value = "Additional"
            ws.cell(insert_row, 1).font = Font(bold=True)
            insert_row += 1
            need_header = False

        ws.cell(insert_row, 1).value = ticker
        ws.cell(insert_row, 2).value = bbg
        src_map[insert_row] = {'ticker': ticker, 'bbg': bbg}
        added.append({'row': insert_row, 'ticker': ticker, 'bbg': bbg})
        insert_row += 1
        print(f"ADD {ticker} at row {insert_row - 1}")

    # Pull data for new tickers
    if added:
        print(f"\nPulling data for {len(added)} new tickers...")
        bbg_list = [a['bbg'] for a in added]
        snap = bdp_dict(bbg_list, ['PX_LAST','CUR_MKT_CAP','CHG_PCT_YTD','NET_DEBT_TO_EBITDA',
            'RETURN_ON_INV_CAPITAL','SHORT_AND_LONG_TERM_DEBT','BS_CASH_NEAR_CASH_ITEM',
            'TOTAL_EQUITY','TRAIL_12M_OPER_INC','TRAIL_12M_EBITDA'])

        eps, rev, ebit = {}, {}, {}
        for yr in ['24','25','26','27','28']:
            for field, store, key in [('BEST_EPS',eps,'EPS'),('BEST_SALES',rev,'REV'),('BEST_EBIT',ebit,'EBIT')]:
                d = bdp_ov(bbg_list, field, f'{yr}Y')
                for t in d:
                    if t not in store: store[t] = {}
                    store[t][f'{key}_20{yr}'] = d[t].get(field)

        for a in added:
            row, bbg = a['row'], a['bbg']
            s = snap.get(bbg, {})
            e = eps.get(bbg, {})
            r = rev.get(bbg, {})
            eb = ebit.get(bbg, {})
            price = s.get('PX_LAST')
            mktcap = s.get('CUR_MKT_CAP')
            if mktcap: mktcap /= 1e9
            if price: ws.cell(row, 5).value = price
            if mktcap: ws.cell(row, 4).value = mktcap
            for ci, yr in enumerate(['2026','2027','2028']):
                ev = e.get(f'EPS_{yr}')
                if price and ev and ev != 0: ws.cell(row, 6+ci).value = price/ev
            net_levg = s.get('NET_DEBT_TO_EBITDA')
            if net_levg: ws.cell(row, 9).value = net_levg
            rev_26 = r.get('REV_2026')
            ebit_26 = eb.get('EBIT_2026')
            if rev_26: ws.cell(row, 10).value = rev_26/1000
            if rev_26 and ebit_26 and rev_26 != 0: ws.cell(row, 11).value = ebit_26/rev_26
            roic = s.get('RETURN_ON_INV_CAPITAL')
            if roic: ws.cell(row, 12).value = roic/100
            def cagr(start, end, yrs=3):
                if not start or not end or start <= 0 or end <= 0: return None
                return (end/start)**(1/yrs)-1
            v = cagr(r.get('REV_2025'), r.get('REV_2028'))
            if v: ws.cell(row, 13).value = v
            v = cagr(eb.get('EBIT_2025'), eb.get('EBIT_2028'))
            if v: ws.cell(row, 14).value = v
            v = cagr(e.get('EPS_2025'), e.get('EPS_2028'))
            if v: ws.cell(row, 15).value = v
            ytd = s.get('CHG_PCT_YTD')
            if ytd: ws.cell(row, 17).value = ytd/100
            for ci, yr in enumerate(['2024','2025','2026','2027','2028']):
                v = e.get(f'EPS_{yr}')
                if v: ws.cell(row, 36+ci).value = v
                v = r.get(f'REV_{yr}')
                if v: ws.cell(row, 42+ci).value = v
                v = eb.get(f'EBIT_{yr}')
                if v: ws.cell(row, 48+ci).value = v
            debt = s.get('SHORT_AND_LONG_TERM_DEBT', 0) or 0
            cash = s.get('BS_CASH_NEAR_CASH_ITEM', 0) or 0
            if mktcap:
                ws.cell(row, 56).value = mktcap
                ws.cell(row, 55).value = debt - cash
                ws.cell(row, 54).value = mktcap + debt - cash
            ltm_ebitda = s.get('TRAIL_12M_EBITDA')
            if ltm_ebitda: ws.cell(row, 57).value = ltm_ebitda
            print(f"  {a['ticker']}: ${price}")

    # --- Rebuild BBG Live with FYE-adjusted formulas ---
    print("\nRebuilding BBG Live...")
    if 'BBG Live' in wb.sheetnames:
        del wb['BBG Live']
    live = wb.copy_worksheet(ws)
    live.title = "BBG Live"

    # Pull FYE
    all_bbg = list(set(info['bbg'] for info in src_map.values()))
    print(f"Pulling FYE for {len(all_bbg)} tickers...")
    fye_data = bdp_dict(all_bbg, ['EQY_FISCAL_YR_END'])
    # Flatten: fye_data[bbg] = {'EQY_FISCAL_YR_END': 'MM/YYYY'}
    fye_flat = {}
    for bbg, fields in fye_data.items():
        fye_flat[bbg] = fields.get('EQY_FISCAL_YR_END', '')

    blue = Font(color="0000FF")
    live.cell(2, 3).value = 'FYE'
    live.cell(1, 6).value = 'PE'
    live.cell(2, 6).value = 2026
    live.cell(2, 7).value = 2027
    live.cell(2, 8).value = 2028
    live.cell(1, 10).value = '2026 estimates'
    live.cell(1, 13).value = 'Growth: 2025-2028'
    live.cell(2, 20).value = "25-'28 EPS gwth"

    shifted = []
    for row, info in src_map.items():
        bbg = info['bbg']
        fye_label, offset = get_fye_offset(bbg, fye_flat)
        live.cell(row, 3).value = fye_label

        def fy(cy): return f'{(cy+offset)%100}Y'
        def bdp(fld, ov=None):
            if ov: return f'=BDP("{bbg}","{fld}","BEST_FPERIOD_OVERRIDE","{ov}")'
            return f'=BDP("{bbg}","{fld}")'

        pr = f'$E${row}'
        live.cell(row, 4).value = f'={bdp("CUR_MKT_CAP")[1:]}/1000000000'
        live.cell(row, 5).value = bdp("PX_LAST")
        for ci, cy in enumerate([2026,2027,2028]):
            live.cell(row,6+ci).value = f'=IF({bdp("BEST_EPS",fy(cy))[1:]}<>0,{pr}/{bdp("BEST_EPS",fy(cy))[1:]},"")'
        live.cell(row, 9).value = bdp("NET_DEBT_TO_EBITDA")
        live.cell(row, 10).value = f'={bdp("BEST_SALES",fy(2026))[1:]}/1000'
        live.cell(row, 11).value = f'=IF({bdp("BEST_SALES",fy(2026))[1:]}<>0,{bdp("BEST_EBIT",fy(2026))[1:]}/{bdp("BEST_SALES",fy(2026))[1:]},"")'
        live.cell(row, 12).value = f'={bdp("RETURN_ON_INV_CAPITAL")[1:]}/100'
        for col, fld in [(13,'BEST_SALES'),(14,'BEST_EBIT'),(15,'BEST_EPS')]:
            live.cell(row, col).value = f'=IF(AND({bdp(fld,fy(2025))[1:]}>0,{bdp(fld,fy(2028))[1:]}>0),({bdp(fld,fy(2028))[1:]}/{bdp(fld,fy(2025))[1:]})^(1/3)-1,"")'
        live.cell(row, 17).value = f'={bdp("CHG_PCT_YTD")[1:]}/100'
        live.cell(row, 20).value = f'=IF(AND(F{row}<>"",O{row}<>"",O{row}<>0),F{row}/(O{row}*100),"")'
        live.cell(row, 28).value = bdp("TRAIL_12M_OPER_INC")
        live.cell(row, 29).value = bdp("SHORT_AND_LONG_TERM_DEBT")
        live.cell(row, 30).value = bdp("TOTAL_EQUITY")
        live.cell(row, 31).value = bdp("BS_CASH_NEAR_CASH_ITEM")
        live.cell(row, 32).value = bdp("BS_LT_INVEST")
        live.cell(row, 33).value = f'=AC{row}+AD{row}'
        live.cell(row, 34).value = f'={bdp("RETURN_ON_INV_CAPITAL")[1:]}/100'
        for ci, cy in enumerate([2024,2025,2026,2027,2028]):
            live.cell(row, 36+ci).value = bdp("BEST_EPS", fy(cy))
            live.cell(row, 42+ci).value = bdp("BEST_SALES", fy(cy))
            live.cell(row, 48+ci).value = bdp("BEST_EBIT", fy(cy))
        live.cell(row, 54).value = f'=D{row}+BC{row}'
        live.cell(row, 55).value = f'=AC{row}-AE{row}'
        live.cell(row, 56).value = f'=D{row}'
        live.cell(row, 57).value = bdp("TRAIL_12M_EBITDA")

        if offset: shifted.append(f"{info['ticker']} ({fye_label})")

    # Blue font
    for row in range(1, live.max_row + 1):
        for col in range(1, live.max_column + 1):
            val = live.cell(row, col).value
            if val and isinstance(val, str) and 'BDP(' in val:
                live.cell(row, col).font = blue

    wb.save(OUT)
    print(f"\nSaved: {OUT}")
    print(f"\nCY-shifted ({len(shifted)}):")
    for s in shifted: print(f"  {s}")


if __name__ == '__main__':
    main()
